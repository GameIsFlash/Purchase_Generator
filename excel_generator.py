import openpyxl
from openpyxl.styles import Border, Side, PatternFill, Font, Alignment
from openpyxl.drawing import image
from pathlib import Path
from datetime import datetime
import tempfile
import logging
from config import ERROR_MESSAGES, TABLE_CONFIG

logger = logging.getLogger(__name__)

class ExcelGenerator:
    def __init__(self, output_dir: str = "output"):
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(exist_ok=True)
        self.temp_dir = Path(tempfile.gettempdir()) / "purchase_tables"
        self.temp_dir.mkdir(exist_ok=True)

    def generate_table(self, supplier_data):
        """Генерация таблицы закупки"""
        try:
            supplier_name = supplier_data['supplier']
            items = supplier_data['items']
            date_str = self._format_date()
            filename = f"Закупка {supplier_name} {date_str}.xlsx"
            filepath = self.output_dir / filename

            if filepath.exists():
                filepath.unlink()

            workbook = openpyxl.Workbook()
            worksheet = workbook.active
            worksheet.title = f"Закупка {supplier_name}"

            self._format_header(worksheet)
            temp_files = self._fill_data(worksheet, items)
            self._create_total_row(worksheet, items)

            workbook.save(str(filepath))
            workbook.close()

            for temp_file in temp_files:
                try:
                    if temp_file.exists():
                        temp_file.unlink()
                except:
                    pass

            logger.info(f"Excel файл создан: {filepath}")
            return str(filepath)

        except Exception as e:
            error_msg = ERROR_MESSAGES['EXCEL_GENERATION_ERROR'].format(str(e))
            logger.error(error_msg)
            raise Exception(error_msg)

    def generate_general_table(self, supplier_data):
        """Генерация общей таблицы всех товаров поставщика"""
        try:
            supplier_name = supplier_data['supplier']
            items = supplier_data['items']
            date_str = self._format_date()
            filename = f"Наличие {supplier_name} {date_str}.xlsx"
            filepath = self.output_dir / filename

            if filepath.exists():
                filepath.unlink()

            workbook = openpyxl.Workbook()
            worksheet = workbook.active
            worksheet.title = f"Наличие {supplier_name}"

            self._format_general_header(worksheet)
            temp_files = self._fill_general_data(worksheet, items)
            self._create_general_total_row(worksheet, items)

            workbook.save(str(filepath))
            workbook.close()

            for temp_file in temp_files:
                try:
                    if temp_file.exists():
                        temp_file.unlink()
                except:
                    pass

            logger.info(f"Общая таблица создана: {filepath}")
            return str(filepath)

        except Exception as e:
            error_msg = ERROR_MESSAGES['EXCEL_GENERATION_ERROR'].format(str(e))
            logger.error(error_msg)
            raise Exception(error_msg)

    def _format_header(self, worksheet):
        """Форматирование заголовка для таблицы закупки"""
        for col_letter, width in TABLE_CONFIG['column_widths'].items():
            worksheet.column_dimensions[col_letter].width = width

        header_fill = PatternFill(start_color=TABLE_CONFIG['header_color'], end_color=TABLE_CONFIG['header_color'], fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')
        border = Border(
            left=Side(border_style='thick', color='000000'),
            right=Side(border_style='thick', color='000000'),
            top=Side(border_style='thick', color='000000'),
            bottom=Side(border_style='thick', color='000000')
        )
        alignment = Alignment(horizontal='center', vertical='center')

        for col_num, header in enumerate(TABLE_CONFIG['headers'], 1):
            cell = worksheet.cell(row=1, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = alignment
            cell.border = border

    def _format_general_header(self, worksheet):
        """Форматирование заголовка для общей таблицы"""
        # Настройки колонок для общей таблицы (без количества и суммы)
        general_column_widths = {'A': 30, 'B': 35, 'C': 12, 'D': 20}
        general_headers = ['Фото', 'Наименование', 'Цена', 'Артикул']

        for col_letter, width in general_column_widths.items():
            worksheet.column_dimensions[col_letter].width = width

        header_fill = PatternFill(start_color='009900', end_color='009900', fill_type='solid')  # Зеленый цвет для общей таблицы
        header_font = Font(bold=True, color='FFFFFF')
        border = Border(
            left=Side(border_style='thick', color='000000'),
            right=Side(border_style='thick', color='000000'),
            top=Side(border_style='thick', color='000000'),
            bottom=Side(border_style='thick', color='000000')
        )
        alignment = Alignment(horizontal='center', vertical='center')

        for col_num, header in enumerate(general_headers, 1):
            cell = worksheet.cell(row=1, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = alignment
            cell.border = border

    def _fill_data(self, worksheet, items):
        """Заполнение данных для таблицы закупки"""
        border = Border(
            left=Side(border_style='thick', color='000000'),
            right=Side(border_style='thick', color='000000'),
            top=Side(border_style='thick', color='000000'),
            bottom=Side(border_style='thick', color='000000')
        )
        center_align = Alignment(horizontal='center', vertical='center')
        left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)

        temp_files = []

        for row_num, item in enumerate(items, 2):
            worksheet.row_dimensions[row_num].height = TABLE_CONFIG['row_height']

            if item.get('processed_image'):
                image_file = self._save_temp_image(item['processed_image'], row_num)
                if image_file:
                    try:
                        excel_img = image.Image(str(image_file))
                        excel_img.anchor = f'A{row_num}'
                        worksheet.add_image(excel_img)
                        temp_files.append(image_file)
                    except Exception as e:
                        logger.warning(f"Ошибка вставки изображения: {e}")

            self._fill_row_cells(worksheet, row_num, item, border, center_align, left_align)

        return temp_files

    def _fill_general_data(self, worksheet, items):
        """Заполнение данных для общей таблицы"""
        border = Border(
            left=Side(border_style='thick', color='000000'),
            right=Side(border_style='thick', color='000000'),
            top=Side(border_style='thick', color='000000'),
            bottom=Side(border_style='thick', color='000000')
        )
        center_align = Alignment(horizontal='center', vertical='center')
        left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)

        temp_files = []

        for row_num, item in enumerate(items, 2):
            worksheet.row_dimensions[row_num].height = TABLE_CONFIG['row_height']

            # Изображение
            if item.get('processed_image'):
                image_file = self._save_temp_image(item['processed_image'], row_num)
                if image_file:
                    try:
                        excel_img = image.Image(str(image_file))
                        excel_img.anchor = f'A{row_num}'
                        worksheet.add_image(excel_img)
                        temp_files.append(image_file)
                    except Exception as e:
                        logger.warning(f"Ошибка вставки изображения: {e}")

            # Заполнение ячеек для общей таблицы
            worksheet.cell(row=row_num, column=1).border = border

            # Наименование
            cell_b = worksheet.cell(row=row_num, column=2, value=item['name'])
            cell_b.border = border
            cell_b.alignment = center_align

            # Цена
            cell_c = worksheet.cell(row=row_num, column=3, value=float(item['price']))
            cell_c.border = border
            cell_c.alignment = center_align
            cell_c.number_format = '#,##0.00'

            # Артикул
            cell_d = worksheet.cell(row=row_num, column=4, value=item['article'])
            cell_d.border = border
            cell_d.alignment = center_align

        return temp_files

    def _fill_row_cells(self, worksheet, row_num, item, border, center_align, left_align):
        """Заполнение ячеек строки для таблицы закупки"""
        worksheet.cell(row=row_num, column=1).border = border

        cell_b = worksheet.cell(row=row_num, column=2, value=item['name'])
        cell_b.border = border
        cell_b.alignment = center_align

        cell_c = worksheet.cell(row=row_num, column=3, value=float(item['price']))
        cell_c.border = border
        cell_c.alignment = center_align
        cell_c.number_format = '#,##0.00'

        cell_d = worksheet.cell(row=row_num, column=4, value=item['quantity'])
        cell_d.border = border
        cell_d.alignment = center_align

        total_amount = float(item['price']) * item['quantity']
        cell_e = worksheet.cell(row=row_num, column=5, value=total_amount)
        cell_e.border = border
        cell_e.alignment = center_align
        cell_e.number_format = '#,##0.00'

    def _create_total_row(self, worksheet, items):
        """Создание итоговой строки для таблицы закупки"""
        total_row = len(items) + 2
        worksheet.merge_cells(f'A{total_row}:C{total_row}')

        border = Border(
            left=Side(border_style='thick', color='000000'),
            right=Side(border_style='thick', color='000000'),
            top=Side(border_style='thick', color='000000'),
            bottom=Side(border_style='thick', color='000000')
        )
        center_align = Alignment(horizontal='center', vertical='center')
        bold_font = Font(bold=True)

        total_cell = worksheet.cell(row=total_row, column=1, value=f"Итого: {len(items)} позиций")
        total_cell.font = bold_font
        total_cell.alignment = center_align
        total_cell.border = border

        for col in [2, 3]:
            worksheet.cell(row=total_row, column=col).border = border

        quantity_cell = worksheet.cell(row=total_row, column=4, value=sum(item['quantity'] for item in items))
        quantity_cell.font = bold_font
        quantity_cell.border = border
        quantity_cell.alignment = center_align

        total_amount = sum(float(item['price']) * item['quantity'] for item in items)
        amount_cell = worksheet.cell(row=total_row, column=5, value=total_amount)
        amount_cell.font = bold_font
        amount_cell.border = border
        amount_cell.alignment = center_align
        amount_cell.number_format = '#,##0.00'

    def _create_general_total_row(self, worksheet, items):
        """Создание итоговой строки для общей таблицы"""
        total_row = len(items) + 2
        worksheet.merge_cells(f'A{total_row}:D{total_row}')

        border = Border(
            left=Side(border_style='thick', color='000000'),
            right=Side(border_style='thick', color='000000'),
            top=Side(border_style='thick', color='000000'),
            bottom=Side(border_style='thick', color='000000')
        )
        center_align = Alignment(horizontal='center', vertical='center')
        bold_font = Font(bold=True)

        total_cell = worksheet.cell(row=total_row, column=1, value=f"Всего позиций: {len(items)}")
        total_cell.font = bold_font
        total_cell.alignment = center_align
        total_cell.border = border

        for col in [2, 3, 4]:
            worksheet.cell(row=total_row, column=col).border = border

    def _save_temp_image(self, processed_image, row_num):
        """Сохранение временного изображения"""
        try:
            temp_file = self.temp_dir / f"temp_{row_num}.png"
            processed_image.save(str(temp_file), 'PNG')
            return temp_file
        except Exception as e:
            logger.error(f"Ошибка сохранения временного изображения: {e}")
            return None

    def _format_date(self):
        """Форматирование даты"""
        now = datetime.now()
        months = {
            1: "января", 2: "февраля", 3: "марта", 4: "апреля",
            5: "мая", 6: "июня", 7: "июля", 8: "августа",
            9: "сентября", 10: "октября", 11: "ноября", 12: "декабря"
        }
        return f"{now.day:02d} {months[now.month]} {now.year % 100:02d}"